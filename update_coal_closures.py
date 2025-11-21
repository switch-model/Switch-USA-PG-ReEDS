from pathlib import Path
import openpyxl
import pandas as pd
import numpy as np
import re
from scipy.optimize import linprog
from powergenome.generators import download_860m
from powergenome.util import load_settings
from powergenome.params import DATA_PATHS

settings_dir = "pg/settings"


eia_skiprows = 2
eia_skipfooter = 2
eia_sheet_name = "Operating"

settings = load_settings(settings_dir)
eia_860m_fn = settings.get("eia_860m_fn")
if not eia_860m_fn:
    raise ValueError(f"No `eia_860m_fn` parameter defined in {settings_dir}")

# GEM_coal_tracker_url = settings.get("GEM_coal_tracker_url")
GEM_coal_tracker_file = (
    Path(settings["input_folder"]) / "Global-Coal-Plant-Tracker-July-2025.xlsx"
)

# Force re-download of the EIA workbook (will also force re-creation of
# the cached "operating" worksheet next time PowerGenome runs)
eia_wb_path = DATA_PATHS["eia_860m"] / eia_860m_fn
if eia_wb_path.exists():
    eia_wb_path.unlink()
download_860m(eia_860m_fn)  # re-create file at wb_path

# read EIA data using settings for ~2025 workbooks
eia = pd.read_excel(
    eia_wb_path,
    sheet_name=eia_sheet_name,
    na_values=[" ", ""],
    keep_default_na=False,
    skiprows=eia_skiprows,
    skipfooter=eia_skipfooter,
)
# tag each generator with the corresponding row in the Excel sheet
# (1-based to match the view in Excel)
eia["eia_row"] = eia.index + eia_skiprows + 2  # 0->1 and skip header
# only keep coal plants
eia = eia[eia["Technology"].str.contains("Coal")]
# GEM only covers plants >= 30 MW
eia = eia[eia["Nameplate Capacity (MW)"] >= 30]

# get data from GEM coal closures workbook
gem = pd.read_excel(GEM_coal_tracker_file, sheet_name="Units")
gem["gem_row"] = gem.index + 2  # 0->1 and skip header
gem = gem[gem["Country/Area"] == "United States"].copy()
# strip out extra "timepoint 1" notations
gem["Unit name"] = gem["Unit name"].str.replace(", timepoint 1", "")
# treat actual retirements as planned retirements (some of these are not in EIA)
gem["Planned retirement"] = gem["Planned retirement"].fillna(gem["Retired year"])

# open EIA 860m workbook and GEM coal closures workbook,
# then calculate the "distance" from each generator in 860m to each generator in
# GEM, based on
# candidates:
# - spatial distance < 3 miles (boolean)
# - size match < 1 (boolean)
# score:
# - first word of plant name matches (boolean)
# - unit digits match (boolean)
# - unit digits blank (boolean)
# - size difference
# - start year difference


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0088  # mean Earth radius [km]
    φ1, λ1, φ2, λ2 = map(np.deg2rad, [lat1, lon1, lat2, lon2])
    dφ = φ2 - φ1
    dλ = λ2 - λ1
    a = np.sin(dφ / 2) ** 2 + np.cos(φ1) * np.cos(φ2) * np.sin(dλ / 2) ** 2
    return 2 * R * np.arcsin(np.sqrt(a))


# calculate "distance" from each eia row to each gem row
distances = []
for i, e in eia.iterrows():
    # Only consider gem rows within 3 km
    g = gem[
        haversine_km(e["Latitude"], e["Longitude"], gem["Latitude"], gem["Longitude"])
        < 3
    ]
    # only consider plants within 25% of same size
    g = g[(g["Capacity (MW)"] / e["Nameplate Capacity (MW)"] - 1).abs() <= 0.25]
    # calculate weighted "distance" score
    d = 0
    # first word of plant name differs
    d = 10 * d + (g["Plant name"].str.split().str[0] != e["Plant Name"].split()[0])
    # unit digits differ
    d = 10 * d + (
        g["Unit name"].str.replace(r"\D+", "", regex=True)
        != re.sub(r"\D+", "", e["Generator ID"])
    )
    # size difference (MW)
    d = 1000 * d + (g["Capacity (MW)"] - e["Nameplate Capacity (MW)"]).abs()
    # start year difference (1 year treated as equivalent to 1 MW); sometimes missing for cancelled projects
    d = d + (g["Start year"] - e["Operating Year"]).abs().fillna(100)
    distances.extend((e["eia_row"], r_, d_) for d_, r_ in zip(d, g["gem_row"]))
    if e["eia_row"] == 13369:
        break

# Use a linear program to
# assign eia rows to gem rows, minimizing a distance score
# This prioritizes assigning each row to a partner, and then
# secondarily minimizing the sum of the distances between
# the partners.
penalty = 100000  # penalty for leaving rows unmatched
eia_rows = sorted({i for i, _, _ in distances})
gem_rows = sorted({j for _, j, _ in distances})

# Variables: [w_ij for all listed pairs] + [s_i for each eia id] + [t_j for each gem id]
pairs = [(i, j) for i, j, _ in distances]
nW, nE, nG = len(pairs), len(eia_rows), len(gem_rows)

# Objective: minimize sum w_ij * dist_ij + penalty * (sum s_i + sum t_j)
c_w = np.array([d for _, _, d in distances], dtype=float)
c_s = np.full(nE, float(penalty))
c_t = np.full(nG, float(penalty))
c = np.concatenate([c_w, c_s, c_t])

# Equality constraints:
# For each eia i: sum_j w_ij + s_i = 1
# For each gem j: sum_i w_ij + t_j = 1
A_eq = np.zeros((nE + nG, nW + nE + nG), dtype=float)
b_eq = np.ones(nE + nG, dtype=float)

eia_pos = {i: k for k, i in enumerate(eia_rows)}
gem_pos = {j: k for k, j in enumerate(gem_rows)}

# Rows for eia constraints
for k, (i, j) in enumerate(pairs):
    A_eq[eia_pos[i], k] += 1.0
# Add s_i coefficients (identity)
for i, r in eia_pos.items():
    A_eq[r, nW + r] = 1.0

# Rows for gem constraints
base = nE
for k, (i, j) in enumerate(pairs):
    A_eq[base + gem_pos[j], k] += 1.0
# Add t_j coefficients (identity)
for j, r in gem_pos.items():
    A_eq[base + r, nW + nE + r] = 1.0

# Bounds:
# 0 <= w_ij <= 1; 0 <= s_i <= 1; 0 <= t_j <= 1
bounds = [(0.0, 1.0)] * nW + [(0.0, 1.0)] * nE + [(0.0, 1.0)] * nG

res = linprog(c, A_eq=A_eq, b_eq=b_eq, bounds=bounds, method="highs")
if not res.success:
    raise RuntimeError(f"LP failed: {res.message}")

x = res.x
w = x[:nW]
s = x[nW : nW + nE]
t = x[nW + nE :]

weights = {pair: w[idx] for idx, pair in enumerate(pairs)}
eia_shortfall = {i: s[eia_pos[i]] for i in eia_rows if s[eia_pos[i]] > 0}
gem_shortfall = {j: t[gem_pos[j]] for j in gem_rows if t[gem_pos[j]] > 0}

# make sure we got exact 1-1 matching
assert all(w in {0, 1} for w in weights.values())

# make dicts showing the winning matches
gem_for_eia = {e: g for (e, g), w in weights.items() if w == 1}
eia_for_gem = {g: e for (e, g), w in weights.items() if w == 1}

# report any that didn't match (for EIA it's one oddball one; for GEM, they
# all seem to be cases where the EIA plant is now marked as natural gas, not coal)
missing_candidates = set(eia["eia_row"]) - set(eia_rows)
if eia_shortfall or missing_candidates:
    print(f"Unmatched EIA rows:")
    for (e, g), d in zip(pairs, c_w):
        if e in eia_shortfall:
            alt_match = f" (-> eia {eia_for_gem[g]})" if g in eia_for_gem else ""
            print(f"eia {e}, gem {g}{alt_match}: dist={d}, weight={weights[e, g]}")
    for e in missing_candidates:
        print(f"eia {e}: no candidates in gem")

gem_operating = set(gem[gem["Status"] == "operating"]["gem_row"])
missing_candidates = gem_operating - set(gem_rows)
if gem_shortfall or missing_candidates:
    print(f"Unmatched GEM rows:")
    for (e, g), d in sorted(zip(pairs, c_w), key=lambda x: (x[0][1], x[0][0])):
        if g in gem_shortfall and g in gem_operating:
            alt_match = f" (-> gem {gem_for_eia[e]})" if e in gem_for_eia else ""
            print(f"gem {g}, eia {e}{alt_match}: dist={d}, weight={weights[e, g]}")
    for g in missing_candidates:
        print(f"gem {g}: no candidates in eia")

# check that status for matched GEM projects is online or retired
assert set(gem[gem["gem_row"].isin(eia_for_gem)]["Status"].unique()) == {
    "operating",
    "retired",
}

# calculate new retirement date for each row in EIA table, based on best matching GEM
# warn of any cases where
new_retirement_year = (
    eia["eia_row"].map(gem_for_eia).map(gem.set_index("gem_row")["Planned retirement"])
)

# apply new retirement year unless it is nan (sometimes they disagree, but we assume
# gem is more up to date); then convert remaining NaNs to " " to match EIA workbook
eia["new_retirement_year"] = new_retirement_year.fillna(eia["Planned Retirement Year"])

# save results back to EIA 860 workbook
print(f"Opening {eia_wb_path}.")
eia_wb = openpyxl.load_workbook(eia_wb_path)
print(f"Updating workbook.")
eia_ws = eia_wb["Operating"]
month_col = eia.columns.get_loc("Planned Retirement Month") + 1  # openpyxl is 1-based
year_col = eia.columns.get_loc("Planned Retirement Year") + 1
for i, (row, new_year, old_year) in eia[
    ["eia_row", "new_retirement_year", "Planned Retirement Year"]
].iterrows():
    # only update values that have changed, to simplify inspection of the workbook
    if new_year != old_year and not (np.isnan(new_year) and np.isnan(old_year)):
        if np.isnan(new_year):
            year = " "
            month = " "
        else:
            year = new_year
            month = 6  # assume mid-year
        eia_ws.cell(row=row, column=year_col).value = year
        eia_ws.cell(row=row, column=month_col).value = month

print(f"Saving workbook.")
eia_wb.save(eia_wb_path)
print(f"Finished updating {eia_wb_path}.")
